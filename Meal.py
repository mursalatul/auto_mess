import csv
from openpyxl import Workbook, load_workbook
import calendar
from datetime import datetime, timedelta, date


class Meal:
    def __init__(self, filec) -> None:
        self.xl_file = filec

    def getDaysInMonth(self, year: int, month: int):
        """
        return number of days in the following month of the year
        """
        last_day = calendar.monthrange(year, month)[1]
        return last_day

    def initializeSheet(self, names: list):
        """
        this method will only called first day of the month to create and initialize the sheet.
        previous date will be removed as new sheet will be created
        """
        # adding dates in the sheet creating date
        present_day = datetime.now().strftime("%d-%m-%Y")
        year = date.today().year
        month = date.today().month
        num_of_days_in_month = self.getDaysInMonth(year, month)

        # creating workbook and sheet
        # workbook = load_workbook(self.xl_file)
        workbook = Workbook()
        sheet1 = workbook.active

        # adding dates in the 1st column
        for i in range(0, num_of_days_in_month):
            next_day = (datetime.now() + timedelta(days=i)
                        ).strftime("%d-%m-%Y")
            cell = sheet1.cell(row=i + 2, column=1)
            cell.value = next_day

        # adding names in the 1st row
        for i, ele in enumerate(names):
            cell = sheet1.cell(row=1, column=i + 2)
            cell.value = ele
        # saving the sheet
        workbook.save(filename="mealdataxx.xlsx")

    def readMeal(self, date: str, name: str = 'All') -> dict:
        """
        return meal of a specific date.
        Args:
            date(DD-MM-YYYY) : date of meal
            name(optional) : name of a particular participant
        Return:
            dict    : name with number of meal
        """
        wb = load_workbook(self.xl_file)
        sheet1 = wb.active
        all_meals = {
            'Adil': 0,
            'Elias': 0,
            'Labib': 0,
            'Nahid': 0,
            'Nurul': 0,
            'Pallob': 0,
            'Prottus': 0,
            'Swadhin': 0
        }

        # iterating the sheet for the specific date row
        date_row = 0
        for i in range(1, 32):
            cell = sheet1.cell(row=i, column=1)
            if cell.value == date:
                date_row = i

        # getting all the meals from that that row and saving in the dict
        for meal_number in range(2, 10):
            cell = sheet1.cell(row=1, column=meal_number)
            cell_head = cell.value
            cell = sheet1.cell(row=date_row, column=meal_number)
            all_meals[cell_head] = cell.value

        # if no name provided, return all meal, else particular name meal
        if name == 'All':
            return all_meals
        else:
            return {name: all_meals[name]}

    def writeAMeal(self, date: str, name: str, total_meal: int) -> None:
        """
        Write meal for a particular person
        Args:
            date(DD-MM-YYYY)    : date of the meal
            name    : meal holder name
        Return:
            None
        """
        wb = load_workbook(self.xl_file)
        sheet1 = wb.active

        # finding the row of date
        # iterating the sheet for the specific date row
        date_row = 0
        for i in range(1, 32):
            cell = sheet1.cell(row=i, column=1)
            if cell.value == date:
                date_row = i
        # finding the name culumn
        name_col
        for c in range (2, 10):
            cell = sheet1.cell(row=1, column=c):
            if cell.value == name:
                name_col = c

        # setting the meal
        sheet1.cell(row=date_row, column=name_col).value = total_meal
        wb.save(self.xl_file)

    # def writeAllMeal(self):
    
# x = Meal('mealdata.xlsx')
# x.initializeSheet(sorted(
#     ["Adil", "Elias", "Nurul", "Labib", "Prottus", "Pallob", "Nahid", "Swadhin"]))
