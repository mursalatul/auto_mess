from openpyxl import Workbook, load_workbook
import calendar
from datetime import datetime, timedelta, date
import os
from reportlab.platypus import SimpleDocTemplate, Table
from botdata import XL_FILE_PATH, SOLO_XL_FILE_PATH, SOLO_XL_FILE


class Meal:
    def __init__(self, filec: str) -> None:
        # targeted xlsx file, store meal data(date, meal of a person)
        self.xl_file = filec

    async def __xl2pdf(self, date: datetime):
        """
        save the sheet as pdf
        Args:
            None
        Return:
            None
        """
        wb = load_workbook(self.xl_file)
        sheet = wb.active

        # creating the pdf
        pdf_filename = f"{XL_FILE_PATH}/Month{date.month}.pdf"

        # Convert sheet data to a list of lists
        data = []
        for row in sheet.iter_rows(values_only=True):
            data.append(row)

        # Create the PDF document and add the table
        doc = SimpleDocTemplate(pdf_filename)
        table = Table(data)

        # Build the PDF document
        elements = [table]
        doc.build(elements)


    async def getDaysInMonth(self, year: int, month: int):
        """
        return number of days in the following month of the year
        """
        last_day = calendar.monthrange(year, month)[1]
        return last_day

    async def initializeSheet(self, names: list):
        """
        this method will only called first day of the month to create and initialize the sheet.
        previous date will be removed as new sheet will be created
        """
        # adding dates in the sheet creating date
        present_day = datetime.now()

        # only initialize new sheet at the first day of the month
        if present_day.day == 1:
            # if there is already a xl_file in the directory then we have to save the previous sheet as pdf
            for filename in os.listdir(SOLO_XL_FILE_PATH): # 
                if filename == SOLO_XL_FILE:
                    await self.__xl2pdf(present_day - timedelta(1))
                    break

            year = date.today().year
            month = date.today().month
            num_of_days_in_month = await self.getDaysInMonth(year, month)

            # creating workbook and sheet
            # workbook = load_workbook(self.xl_file)
            workbook = Workbook()
            sheet1 = workbook.active

            # adding dates in the 1st column
            for i in range(0, num_of_days_in_month):
                next_day = (datetime.now() + timedelta(days=i)
                            ).strftime("%d-%m-%Y")
                cell = sheet1.cell(row=i + 2, column=1)
                cell.value = next_day

            # adding names in the 1st row
            for i, ele in enumerate(names):
                cell = sheet1.cell(row=1, column=i + 2)
                cell.value = ele

            # in the beginning set all the members meal into 11 for rest of the days in the month
            for nm in names:
                await self.writeAllMeal(present_day, nm, '11')
            # saving the sheet
            workbook.save(filename=self.xl_file)

    async def readMeal(self, date: datetime, name: str = 'All') -> dict:
        """
        return meal of a specific date.
        Args:
            date(DD-MM-YYYY) : date of meal
            name(optional) : name of a particular participant
        Return:
            dict    : name with number of meal
        """
        wb = load_workbook(self.xl_file)
        sheet1 = wb.active
        all_meals = {
            'Adil': 0,
            'Elias': 0,
            'Labib': 0,
            'Nahid': 0,
            'Nurul': 0,
            'Pallob': 0,
            'Prottus': 0,
            'Swadhin': 0
        }

        # calculating date_row from inputed date
        date_row = date.day + 1

        # converting datetime -> str
        date = date.strftime('%d-%m-%Y')

        # getting all the meals from that that row and saving in the dict
        for meal_number in range(2, 10):
            cell = sheet1.cell(row=1, column=meal_number)
            cell_head = cell.value  # cell_head represent the name belong to the cell
            cell = sheet1.cell(row=date_row, column=meal_number)
            # assigning the meal info in the dict from xlsx file
            all_meals[cell_head] = str(cell.value)

        # if no name provided, return all meal, else particular name meal
        if name == 'All':
            return all_meals
        else:
            return {name: all_meals[name]}

    async def writeAMeal(self, date: datetime, name: str, total_meal: str) -> None:
        """
        Write meal for a particular person
        Args:
            date(datetime object)    : date of the meal
            name    : meal holder name
            total_meal : number of meals(meal can be 00, 11, 01, 10)
        Return:
            None
        """
        # edit meal from current date to future dates. not previous dates
        present_day = datetime.today().date()
        if (date >= present_day):
            wb = load_workbook(self.xl_file)
            sheet1 = wb.active
            # finding the row of date
            date_row = date.day + 1
            date = date.strftime("%d-%m-%Y")  # converting datetime -> str

            # finding the name culumn
            name_col = 0
            for c in range(2, 10):
                cell = sheet1.cell(row=1, column=c)
                if cell.value == name:
                    name_col = c
                    break
            # setting the meal
            sheet1.cell(row=date_row, column=name_col).value = str(total_meal)
            wb.save(self.xl_file)

    async def writeAllMeal(self, date: datetime, name: str, total_meal: str):
        """
        write all the meal from 'date' to last date of the into total_meal
        Args:
            date(datetime object)    : date of the meal
            name    : meal holder name
            total_meal : number of meals(meal can be '00', '11', '01', '10')
        Return:
            None
        """
        present_day = datetime.today().date()
        # write meal only for today and next days. not previous days
        if (date >= present_day):
            total_days = await self.getDaysInMonth(date.year, date.month)
            reminding_days = total_days - date.day
            for i in range(reminding_days + 1):
                await self.writeAMeal(date + timedelta(i), name, total_meal)
